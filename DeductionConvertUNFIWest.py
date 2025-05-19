import pandas as pd
import numpy as np
import pdfplumber
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from IPython.display import display, HTML
import ipywidgets as widgets
from ipywidgets import FileUpload, Button, Output, HBox, VBox, Label

def extract_text_from_pdf(pdf_path):
    """Extract all text from a PDF file."""
    all_text = []
    
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    all_text.append(text)
        
        return "\n".join(all_text)
    except Exception as e:
        print(f"Error extracting text from PDF: {e}")
        return ""

def parse_pdf_content(content):
    """Parse the PDF content into structured data."""
    lines = content.split('\n')
    
    # Extract report title and week ending
    title = next((line for line in lines if "HARMLESS HARVEST" in line), "HARMLESS HARVEST")
    week_ending = next((line for line in lines if "Week ending" in line), "")
    
    # Initialize data structures
    main_data = []
    headers = ["Brand", "Product", "Unit", "Description", "Invoice", 
               "Ordered", "Shipped", "Wholesale", "Discount%", "MCB%", "MCB",
               "Customer ID", "Customer Name", "Location"]
    
    main_data.append(headers)
    
    # Parse content line by line
    current_location = ""
    current_customer_id = ""
    current_customer_name = ""
    
    for i, line in enumerate(lines):
        line = line.strip()
        
        # Skip empty lines
        if not line:
            continue
        
        # Check for location headers (City followed by state code)
        if re.match(r'^[A-Za-z]+\s+[A-Z]{2}$', line) and "Customer" not in line:
            current_location = line
            continue
        
        # Check for customer information
        if line.startswith("Customer :"):
            customer_match = re.search(r'Customer : \[(\d+)\]-(.*)', line)
            if customer_match:
                current_customer_id = customer_match.group(1)
                current_customer_name = customer_match.group(2).strip()
            continue
        
        # Process product rows (starting with *HRMLSHRVS)
        if line.startswith("*HRMLSHRVS"):
            # Try to parse product line using regex patterns
            try:
                # Split the line into components
                parts = line.split()
                
                # Need at least 11 parts for a valid product line
                if len(parts) < 11:
                    continue
                
                brand = parts[0]
                product = parts[1]
                unit = f"{parts[2]} {parts[3]}"
                
                # Extract description (which may contain spaces)
                desc_index = 4
                description_parts = []
                
                # Find where description ends (usually before invoice number)
                while desc_index < len(parts) and not re.match(r'^\d{6,}$', parts[desc_index]):
                    description_parts.append(parts[desc_index])
                    desc_index += 1
                
                description = " ".join(description_parts)
                
                # If we couldn't find the invoice number, use a different approach
                if desc_index >= len(parts) - 5:
                    # Try alternative pattern matching
                    invoice_match = re.search(r'(\d{8,9})', line)
                    if invoice_match:
                        invoice = invoice_match.group(1)
                        remaining = line[line.index(invoice) + len(invoice):].strip()
                        num_parts = remaining.split()
                        
                        if len(num_parts) >= 5:
                            ordered = int(num_parts[0])
                            shipped = int(num_parts[1])
                            wholesale = float(num_parts[2])
                            discount = num_parts[3]
                            mcb_percent = num_parts[4]
                            mcb = float(num_parts[5]) if len(num_parts) > 5 else 0.0
                        else:
                            continue
                    else:
                        continue
                else:
                    # Normal case where we found the invoice number
                    invoice = parts[desc_index]
                    desc_index += 1
                    
                    # Extract remaining numeric fields
                    ordered = int(parts[desc_index]) if desc_index < len(parts) else 0
                    desc_index += 1
                    
                    shipped = int(parts[desc_index]) if desc_index < len(parts) else 0
                    desc_index += 1
                    
                    wholesale = float(parts[desc_index]) if desc_index < len(parts) else 0.0
                    desc_index += 1
                    
                    discount = parts[desc_index] if desc_index < len(parts) else "0%"
                    desc_index += 1
                    
                    mcb_percent = parts[desc_index] if desc_index < len(parts) else "0%"
                    desc_index += 1
                    
                    mcb = float(parts[desc_index]) if desc_index < len(parts) else 0.0
                
                # Add to main data
                main_data.append([
                    brand, product, unit, description, invoice,
                    ordered, shipped, wholesale, discount, mcb_percent, mcb,
                    current_customer_id, current_customer_name, current_location
                ])
                
            except Exception as e:
                print(f"Error parsing line {i+1}: {line}")
                print(f"Error details: {e}")
    
    return {
        'title': title,
        'week_ending': week_ending,
        'main_data': main_data
    }

def create_summary_tables(main_data):
    """Create summary tables from the main data."""
    # Skip header row
    data_rows = main_data[1:]
    
    # Create DataFrames for easier manipulation
    columns = main_data[0]
    df = pd.DataFrame(data_rows, columns=columns)
    
    # Convert numeric columns
    for col in ['Ordered', 'Shipped']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(int)
    
    for col in ['Wholesale', 'MCB']:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype(float)
    
    # 1. Location Summary
    location_summary = df.groupby('Location').agg({
        'Shipped': 'sum',
        'MCB': 'sum'
    }).reset_index()
    
    location_summary.columns = ['Location', 'Total Items', 'Total MCB']
    location_summary = [location_summary.columns.tolist()] + location_summary.values.tolist()
    
    # 2. Customer Summary
    customer_summary = df.groupby(['Customer ID', 'Customer Name', 'Location']).agg({
        'Shipped': 'sum',
        'MCB': 'sum'
    }).reset_index()
    
    customer_summary.columns = ['Customer ID', 'Customer Name', 'Location', 'Total Items', 'Total MCB']
    customer_summary = [customer_summary.columns.tolist()] + customer_summary.values.tolist()
    
    # 3. Product Summary
    product_summary = df.groupby(['Product', 'Description']).agg({
        'Shipped': 'sum',
        'MCB': 'sum'
    }).reset_index()
    
    product_summary.columns = ['Product', 'Description', 'Total Items', 'Total MCB']
    product_summary = [product_summary.columns.tolist()] + product_summary.values.tolist()
    
    return {
        'location_summary': location_summary,
        'customer_summary': customer_summary,
        'product_summary': product_summary
    }

def save_to_excel(data, output_path):
    """Save all data to an Excel file with multiple sheets."""
    wb = Workbook()
    
    # Create styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    centered_alignment = Alignment(horizontal='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Format currency function
    def format_currency(ws, row, col):
        ws.cell(row=row, column=col).number_format = '$#,##0.00'
    
    # Add Main Data sheet
    ws1 = wb.active
    ws1.title = "Sales Data"
    
    # Add report title and week ending
    ws1.cell(row=1, column=1, value=data['title']).font = Font(bold=True, size=14)
    ws1.cell(row=2, column=1, value=data['week_ending']).font = Font(italic=True)
    
    # Add main data starting at row 4
    start_row = 4
    for r, row in enumerate(data['main_data']):
        for c, cell_value in enumerate(row):
            ws1.cell(row=r+start_row, column=c+1, value=cell_value)
            
            # Apply styles
            if r == 0:  # Header row
                ws1.cell(row=r+start_row, column=c+1).font = header_font
                ws1.cell(row=r+start_row, column=c+1).fill = header_fill
                ws1.cell(row=r+start_row, column=c+1).alignment = centered_alignment
            
            # Format currency columns
            if r > 0 and c in [7, 10]:  # Wholesale and MCB columns
                format_currency(ws1, r+start_row, c+1)
    
    # Apply borders and adjust column widths
    for col in range(1, len(data['main_data'][0]) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 15
        for row in range(start_row, start_row + len(data['main_data'])):
            ws1.cell(row=row, column=col).border = border
    
    # Add Location Summary sheet
    ws2 = wb.create_sheet("Location Summary")
    
    ws2.cell(row=1, column=1, value=data['title']).font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=data['week_ending']).font = Font(italic=True)
    
    start_row = 4
    for r, row in enumerate(data['location_summary']):
        for c, cell_value in enumerate(row):
            ws2.cell(row=r+start_row, column=c+1, value=cell_value)
            
            # Apply styles
            if r == 0:  # Header row
                ws2.cell(row=r+start_row, column=c+1).font = header_font
                ws2.cell(row=r+start_row, column=c+1).fill = header_fill
                ws2.cell(row=r+start_row, column=c+1).alignment = centered_alignment
            
            # Format currency column
            if r > 0 and c == 2:  # Total MCB column
                format_currency(ws2, r+start_row, c+1)
    
    # Apply borders and adjust column widths
    for col in range(1, len(data['location_summary'][0]) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 15
        for row in range(start_row, start_row + len(data['location_summary'])):
            ws2.cell(row=row, column=col).border = border
    
    # Add Customer Summary sheet
    ws3 = wb.create_sheet("Customer Summary")
    
    ws3.cell(row=1, column=1, value=data['title']).font = Font(bold=True, size=14)
    ws3.cell(row=2, column=1, value=data['week_ending']).font = Font(italic=True)
    
    start_row = 4
    for r, row in enumerate(data['customer_summary']):
        for c, cell_value in enumerate(row):
            ws3.cell(row=r+start_row, column=c+1, value=cell_value)
            
            # Apply styles
            if r == 0:  # Header row
                ws3.cell(row=r+start_row, column=c+1).font = header_font
                ws3.cell(row=r+start_row, column=c+1).fill = header_fill
                ws3.cell(row=r+start_row, column=c+1).alignment = centered_alignment
            
            # Format currency column
            if r > 0 and c == 4:  # Total MCB column
                format_currency(ws3, r+start_row, c+1)
    
    # Apply borders and adjust column widths
    for col in range(1, len(data['customer_summary'][0]) + 1):
        col_width = 15
        if col == 2:  # Customer Name column
            col_width = 30
        ws3.column_dimensions[get_column_letter(col)].width = col_width
        
        for row in range(start_row, start_row + len(data['customer_summary'])):
            ws3.cell(row=row, column=col).border = border
    
    # Add Product Summary sheet
    ws4 = wb.create_sheet("Product Summary")
    
    ws4.cell(row=1, column=1, value=data['title']).font = Font(bold=True, size=14)
    ws4.cell(row=2, column=1, value=data['week_ending']).font = Font(italic=True)
    
    start_row = 4
    for r, row in enumerate(data['product_summary']):
        for c, cell_value in enumerate(row):
            ws4.cell(row=r+start_row, column=c+1, value=cell_value)
            
            # Apply styles
            if r == 0:  # Header row
                ws4.cell(row=r+start_row, column=c+1).font = header_font
                ws4.cell(row=r+start_row, column=c+1).fill = header_fill
                ws4.cell(row=r+start_row, column=c+1).alignment = centered_alignment
            
            # Format currency column
            if r > 0 and c == 3:  # Total MCB column
                format_currency(ws4, r+start_row, c+1)
    
    # Apply borders and adjust column widths
    for col in range(1, len(data['product_summary'][0]) + 1):
        col_width = 15
        if col == 2:  # Description column
            col_width = 30
        ws4.column_dimensions[get_column_letter(col)].width = col_width
        
        for row in range(start_row, start_row + len(data['product_summary'])):
            ws4.cell(row=row, column=col).border = border
    
    # Save the workbook
    wb.save(output_path)
    
def convert_pdf_to_excel(pdf_path, output_path=None):
    """
    Convert a PDF with tabular data to Excel format.
    
    Parameters:
    - pdf_path: Path to the PDF file
    - output_path: Optional path for saving the Excel file
    
    Returns:
    - Dictionary containing the extracted data and summary tables
    - Saves Excel file if output_path is provided
    """
    # Extract text from PDF
    print(f"Processing PDF: {pdf_path}")
    all_text = extract_text_from_pdf(pdf_path)
    
    # Parse the extracted text
    data = parse_pdf_content(all_text)
    
    # Create summary tables
    summaries = create_summary_tables(data['main_data'])
    
    # Combine all data
    result = {
        'title': data['title'],
        'week_ending': data['week_ending'],
        'main_data': data['main_data'],
        'location_summary': summaries['location_summary'],
        'customer_summary': summaries['customer_summary'],
        'product_summary': summaries['product_summary']
    }
    
    # Save to Excel if output path is provided
    if output_path:
        save_to_excel(result, output_path)
        print(f"Excel file saved to: {output_path}")
    
    # Display summary of the extraction
    print_extraction_summary(result)
    
    return result

def print_extraction_summary(data):
    """Print a summary of the extracted data."""
    print("\n===== Extraction Summary =====")
    print(f"Report Title: {data['title']}")
    print(f"Week Ending: {data['week_ending']}")
    print(f"Main Data Rows: {len(data['main_data']) - 1}")  # Exclude header
    print(f"Locations: {len(data['location_summary']) - 1}")  # Exclude header
    print(f"Customers: {len(data['customer_summary']) - 1}")  # Exclude header
    print(f"Products: {len(data['product_summary']) - 1}")  # Exclude header
    print("==============================\n")


def process_uploaded_file(file_content, output_filename="converted_excel.xlsx"):
    """
    Process a file uploaded through Jupyter's file upload widget.
    
    Parameters:
    - file_content: The binary content of the uploaded file
    - output_filename: The name of the output Excel file
    
    Returns:
    - Dictionary containing the extracted data
    """
    # Save the uploaded content to a temporary file
    temp_file = "temp_uploaded.pdf"
    with open(temp_file, 'wb') as f:
        f.write(file_content)
    
    # Process the PDF
    result = convert_pdf_to_excel(temp_file, output_filename)
    
    # Clean up the temporary file
    if os.path.exists(temp_file):
        os.remove(temp_file)
    
    return result


#File Upload

# Specify file paths
pdf_file = "Vendor Charge Back 5.10.25.pdf"  # Change this to your PDF file path
excel_output = "output_report.xlsx"  # Change this to your desired output path

# Run the conversion
# Uncomment the following line to execute
#result = convert_pdf_to_excel(pdf_file, excel_output)